import streamlit as st
import pandas as pd
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import io
import base64
import re
from datetime import datetime, timedelta

# ── Page config ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Primebuild Keypay Location Automation",
    page_icon="📋",
    layout="wide",
)

# ── Brand theme ───────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500;600;700&family=Space+Grotesk:wght@400;500;600;700&display=swap');

html, body, [class*="css"] {
    font-family: 'DM Sans', sans-serif;
    background-color: #050505;
    color: #f0f0f0;
}
.stApp { background-color: #050505; }

h1, h2, h3 {
    font-family: 'Space Grotesk', sans-serif;
    background: linear-gradient(90deg, #00c4b4, #0066ff);
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
    background-clip: text;
}

.block-container { padding-top: 2rem; }

.stFileUploader > div {
    background: #0f0f0f;
    border: 1px solid #1a1a2e;
    border-radius: 12px;
}
.stFileUploader label { color: #f0f0f0 !important; }

.stTextInput > div > div > input {
    background: #0f0f0f !important;
    color: #f0f0f0 !important;
    border: 1px solid #1a1a2e !important;
    border-radius: 8px !important;
}
.stTextInput label { color: #f0f0f0 !important; }

div[data-testid="stButton"] > button {
    background: linear-gradient(90deg, #00c4b4, #0066ff);
    color: white;
    font-family: 'Space Grotesk', sans-serif;
    font-weight: 600;
    border: none;
    border-radius: 8px;
    padding: 0.6rem 2rem;
    font-size: 1rem;
    transition: opacity 0.2s;
    width: 100%;
}
div[data-testid="stButton"] > button:hover { opacity: 0.88; }

div[data-testid="stDownloadButton"] > button {
    background: linear-gradient(90deg, #00c4b4, #0066ff);
    color: white;
    font-family: 'Space Grotesk', sans-serif;
    font-weight: 600;
    border: none;
    border-radius: 8px;
    padding: 0.6rem 2rem;
    font-size: 1rem;
    width: 100%;
}

.metric-card {
    background: #0f0f0f;
    border: 1px solid #1a1a2e;
    border-radius: 12px;
    padding: 1.2rem 1.5rem;
    margin-bottom: 0.5rem;
}
.metric-label {
    font-size: 0.75rem;
    color: #888;
    text-transform: uppercase;
    letter-spacing: 0.08em;
    margin-bottom: 0.3rem;
}
.metric-value {
    font-family: 'Space Grotesk', sans-serif;
    font-size: 1.6rem;
    font-weight: 600;
    background: linear-gradient(90deg, #00c4b4, #0066ff);
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
    background-clip: text;
}

.file-tag {
    display: inline-block;
    background: #0f0f0f;
    border: 1px solid #00c4b4;
    border-radius: 6px;
    padding: 0.2rem 0.7rem;
    font-size: 0.8rem;
    color: #00c4b4;
    margin: 0.2rem;
}

.success-box {
    background: #0a1a0f;
    border: 1px solid #00c4b4;
    border-radius: 10px;
    padding: 1rem 1.5rem;
    margin-top: 1rem;
}
.error-box {
    background: #1a0a0a;
    border: 1px solid #ff4444;
    border-radius: 10px;
    padding: 1rem 1.5rem;
    margin-top: 0.5rem;
    color: #ff9999;
}
.divider {
    height: 1px;
    background: linear-gradient(90deg, #00c4b4, #0066ff);
    margin: 1.5rem 0;
    opacity: 0.3;
}
.section-label {
    font-family: 'Space Grotesk', sans-serif;
    font-size: 0.7rem;
    letter-spacing: 0.12em;
    text-transform: uppercase;
    color: #555;
    margin-bottom: 0.4rem;
}
.category-pill {
    display: inline-block;
    border-radius: 20px;
    padding: 0.15rem 0.75rem;
    font-size: 0.78rem;
    font-weight: 600;
    margin: 0.15rem;
}
</style>
""", unsafe_allow_html=True)

# ── Logo ──────────────────────────────────────────────────────────────────────
def get_logo_b64():
    try:
        with open("logo.jpg", "rb") as f:
            return base64.b64encode(f.read()).decode()
    except FileNotFoundError:
        return None

logo_b64 = get_logo_b64()

col_logo, col_title = st.columns([1, 5])
with col_logo:
    if logo_b64:
        st.markdown(f'<img src="data:image/jpeg;base64,{logo_b64}" style="width:80px;margin-top:6px;">', unsafe_allow_html=True)
with col_title:
    st.markdown("## Keypay Location Automation")
    st.markdown('<p style="color:#888;margin-top:-0.5rem;font-size:0.9rem;">Categorise and flag timesheet data from Keypay exports into a structured summary report</p>', unsafe_allow_html=True)

st.markdown('<div class="divider"></div>', unsafe_allow_html=True)

# ── Classification helpers ────────────────────────────────────────────────────

def get_location_prefix(location):
    """Extract the project code prefix letter (C, D, R, N) from a location string."""
    if pd.isna(location) or not isinstance(location, str):
        return None
    parts = location.split('/')
    if len(parts) < 2:
        return None
    mid = parts[1].strip()
    m = re.match(r'^([A-Z])\d{4,}', mid, re.IGNORECASE)
    if m:
        return m.group(1).upper()
    return None

def is_unallocated(location):
    """Return True if location is a home/unallocated location (no project code)."""
    if pd.isna(location) or not isinstance(location, str):
        return True
    return get_location_prefix(location) is None

def is_c_location(location):
    return get_location_prefix(location) == 'C'

def is_non_c_allocated(location):
    prefix = get_location_prefix(location)
    return prefix is not None and prefix != 'C'

def is_self_approved(row):
    """True only when Reviewed By is populated AND matches the employee's full name."""
    reviewed_by = row.get('Reviewed By', '')
    if pd.isna(reviewed_by) or str(reviewed_by).strip() == '':
        return False
    first = str(row.get('First Name', '')).strip()
    surname = str(row.get('Surname', '')).strip()
    full_name = f"{first} {surname}".strip()
    return str(reviewed_by).strip() == full_name

def classify_row(row):
    """
    Returns one of:
      'approved_non_c'
      'unapproved_allocated'
      'self_approved'
      'al_c_costed'
      'exclude'
    Priority order matters — self_approved and al_c_costed are flagged BEFORE
    the main approved_non_c / unapproved_allocated buckets.
    """
    status = str(row.get('Status', '')).strip()
    location = row.get('Location', '')
    work_type = str(row.get('Work Type', '')).strip() if pd.notna(row.get('Work Type')) else ''

    # 1. Processed → exclude
    if status == 'Processed':
        return 'exclude'

    # 2. Self-approved (only for Approved rows with populated Reviewed By)
    if status == 'Approved' and is_self_approved(row):
        return 'self_approved'

    # 3. Approved Annual Leave on a C location → flag
    if status == 'Approved' and work_type == 'Annual Leave Taken' and is_c_location(location):
        return 'al_c_costed'

    # 4. Approved + C location → exclude
    if status == 'Approved' and is_c_location(location):
        return 'exclude'

    # 5. Approved + unallocated/home location → exclude
    if status == 'Approved' and is_unallocated(location):
        return 'exclude'

    # 6. Approved + non-C allocated (D, R, N, etc.) + no special work type → approved_non_c
    if status == 'Approved' and is_non_c_allocated(location) and work_type == '':
        return 'approved_non_c'

    # 6b. Approved + non-C allocated + any other work type → exclude
    if status == 'Approved' and is_non_c_allocated(location) and work_type != '':
        return 'exclude'

    # 7. Submitted + unallocated → unapproved_unallocated (excluded from report in this sample,
    #    kept as separate bucket for future use but not written since N/A shown)
    if status == 'Submitted' and is_unallocated(location):
        return 'unapproved_unallocated'

    # 8. Submitted + allocated (C or non-C) → unapproved_allocated
    if status == 'Submitted' and not is_unallocated(location):
        return 'unapproved_allocated'

    return 'exclude'


# ── Excel builder ─────────────────────────────────────────────────────────────

YELLOW_FILL = PatternFill('solid', fgColor='FFFF00')
BOLD_FONT = Font(bold=True)
HEADER_FONT = Font(bold=True)

COL_WIDTHS = {
    'A': 11.82, 'B': 13.27, 'C': 14.45, 'D': 8.54, 'E': 12.54,
    'F': 10.27, 'G': 58.63, 'H': 10.45, 'I': 11.0,  'J': 10.0,
    'K': 11.0,  'L': 9.09,  'M': 10.55, 'N': 13.82, 'O': 5.45,
    'P': 9.45,  'Q': 27.09, 'R': 15.09, 'S': 12.54, 'T': 17.45,
    'U': 14.45, 'V': 33.82, 'W': 19.09, 'X': 16.55, 'Y': 18.27,
}

def write_category_header(ws, label):
    row_num = ws.max_row + 1
    cell = ws.cell(row=row_num, column=1, value=label)
    cell.fill = YELLOW_FILL
    cell.font = BOLD_FONT

def write_na_row(ws):
    row_num = ws.max_row + 1
    ws.cell(row=row_num, column=1, value='N/A')

def write_data_rows(ws, df_subset, headers):
    for _, row in df_subset.iterrows():
        values = []
        for h in headers:
            v = row.get(h)
            if pd.isna(v):
                v = None
            values.append(v)
        ws.append(values)

def build_output_excel(df_raw, results):
    wb = Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = 'Summary'

    headers = df_raw.columns.tolist()

    # Header row
    ws_summary.append(headers)
    for col_idx, _ in enumerate(headers, 1):
        ws_summary.cell(row=1, column=col_idx).font = HEADER_FONT

    # Set column widths
    for col_letter, width in COL_WIDTHS.items():
        ws_summary.column_dimensions[col_letter].width = width

    # ── Section 1: Approved – Unallocated
    write_category_header(ws_summary, 'Approved – Unallocated')
    if results['approved_unallocated'].empty:
        write_na_row(ws_summary)
    else:
        write_data_rows(ws_summary, results['approved_unallocated'], headers)

    # ── Section 2: Approved – Non-C Locations
    write_category_header(ws_summary, 'Approved – Non-C Locations')
    if results['approved_non_c'].empty:
        write_na_row(ws_summary)
    else:
        write_data_rows(ws_summary, results['approved_non_c'], headers)

    # ── Section 3: Unapproved – Unallocated
    write_category_header(ws_summary, 'Unapproved – Unallocated')
    if results['unapproved_unallocated'].empty:
        write_na_row(ws_summary)
    else:
        write_data_rows(ws_summary, results['unapproved_unallocated'], headers)

    # ── Section 4: Unapproved – Allocated
    write_category_header(ws_summary, 'Unapproved – Allocated')
    if results['unapproved_allocated'].empty:
        write_na_row(ws_summary)
    else:
        write_data_rows(ws_summary, results['unapproved_allocated'], headers)

    # ── Section 5: Others – Self Approved
    write_category_header(ws_summary, 'Others - Self approved timesheets')
    if results['self_approved'].empty:
        write_na_row(ws_summary)
    else:
        write_data_rows(ws_summary, results['self_approved'], headers)

    # ── Section 6: Others – AL C Costed
    write_category_header(ws_summary, "Others - Approved AL but C costed (This is now updated to the employee's HOME location)")
    if results['al_c_costed'].empty:
        write_na_row(ws_summary)
    else:
        write_data_rows(ws_summary, results['al_c_costed'], headers)

    # ── Sheet 2: All Timesheets ───────────────────────────────────────────────
    ws_all = wb.create_sheet('All Timesheets')
    ws_all.append(headers)
    for col_idx, _ in enumerate(headers, 1):
        ws_all.cell(row=1, column=col_idx).font = HEADER_FONT
    for col_letter, width in COL_WIDTHS.items():
        ws_all.column_dimensions[col_letter].width = width
    write_data_rows(ws_all, df_raw, headers)

    # ── Sheet 3: Report Details ───────────────────────────────────────────────
    ws_details = wb.create_sheet('Report Details')
    ws_details.column_dimensions['A'].width = 20
    ws_details.column_dimensions['B'].width = 35
    now_str = datetime.now().strftime('%d/%m/%Y %I:%M\u202f%p').lower()
    now_str = datetime.now().strftime('%d/%m/%Y ') + datetime.now().strftime('%I:%M').lstrip('0') + '\u202f' + datetime.now().strftime('%p').lower()
    ws_details.append(['Report Details', None])
    ws_details.append(['Report Name', 'Timesheets report'])
    ws_details.append(['Date Generated', now_str])
    ws_details.append(['User', ''])
    ws_details.cell(row=1, column=1).font = BOLD_FONT

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


# ── Main UI ───────────────────────────────────────────────────────────────────

st.markdown('<p class="section-label">Step 1 — Upload raw timesheet file</p>', unsafe_allow_html=True)
uploaded_file = st.file_uploader(
    "Upload the Keypay timesheet export (.xlsx)",
    type=["xlsx"],
    label_visibility="collapsed"
)

st.markdown('<div style="margin-top:1rem;"></div>', unsafe_allow_html=True)
st.markdown('<p class="section-label">Step 2 — Output filename</p>', unsafe_allow_html=True)

today_prefix = datetime.now().strftime('%Y%m%d')
default_custom = "Unapproved and Unallocated Timesheets"
custom_name = st.text_input(
    "Custom name (appended after today's date)",
    value=default_custom,
    label_visibility="collapsed",
    placeholder="e.g. Unapproved and Unallocated Timesheets"
)
output_filename = f"{today_prefix}_{custom_name.strip()}.xlsx" if custom_name.strip() else f"{today_prefix}_Unapproved and Unallocated Timesheets.xlsx"
st.markdown(f'<p style="color:#555;font-size:0.8rem;margin-top:0.2rem;">Output: <span style="color:#00c4b4;">{output_filename}</span></p>', unsafe_allow_html=True)

st.markdown('<div class="divider"></div>', unsafe_allow_html=True)

if uploaded_file:
    try:
        df_raw = pd.read_excel(uploaded_file, sheet_name='All Timesheets')
        df_raw = df_raw.dropna(how='all')  # drop fully blank rows

        # ── Classify each row
        df_raw['_category'] = df_raw.apply(classify_row, axis=1)

        results = {
            'approved_non_c':       df_raw[df_raw['_category'] == 'approved_non_c'].drop(columns='_category'),
            'approved_unallocated': df_raw[df_raw['_category'] == 'approved_unallocated'].drop(columns='_category') if 'approved_unallocated' in df_raw['_category'].values else pd.DataFrame(columns=df_raw.columns.drop('_category')),
            'unapproved_unallocated': df_raw[df_raw['_category'] == 'unapproved_unallocated'].drop(columns='_category'),
            'unapproved_allocated': df_raw[df_raw['_category'] == 'unapproved_allocated'].drop(columns='_category'),
            'self_approved':        df_raw[df_raw['_category'] == 'self_approved'].drop(columns='_category'),
            'al_c_costed':          df_raw[df_raw['_category'] == 'al_c_costed'].drop(columns='_category'),
        }

        df_for_output = df_raw.drop(columns='_category')

        # ── Metrics
        total = len(df_raw)
        excluded = len(df_raw[df_raw['_category'] == 'exclude'])
        flagged = len(results['self_approved']) + len(results['al_c_costed'])

        st.markdown("### Results Summary")
        c1, c2, c3, c4, c5, c6 = st.columns(6)
        def metric(col, label, value, color="#00c4b4"):
            col.markdown(f'''
            <div class="metric-card">
                <div class="metric-label">{label}</div>
                <div class="metric-value" style="background:none;-webkit-text-fill-color:{color};">{value}</div>
            </div>''', unsafe_allow_html=True)

        metric(c1, "Total Rows", total)
        metric(c2, "Approved Non-C", len(results['approved_non_c']))
        metric(c3, "Unapproved Allocated", len(results['unapproved_allocated']))
        metric(c4, "Self Approved ⚠️", len(results['self_approved']), "#ffaa00")
        metric(c5, "AL C-Costed ⚠️", len(results['al_c_costed']), "#ffaa00")
        metric(c6, "Excluded", excluded, "#555")

        st.markdown('<div class="divider"></div>', unsafe_allow_html=True)

        # ── Category breakdown table
        st.markdown("### Category Breakdown")
        breakdown_data = {
            "Category": [
                "✅ Approved – Non-C Locations",
                "✅ Approved – Unallocated",
                "⏳ Unapproved – Unallocated",
                "⏳ Unapproved – Allocated",
                "⚠️ Self Approved Timesheets",
                "⚠️ Approved AL but C Costed",
                "🚫 Excluded (Processed / Approved-C / Home)",
            ],
            "Count": [
                len(results['approved_non_c']),
                len(results['approved_unallocated']),
                len(results['unapproved_unallocated']),
                len(results['unapproved_allocated']),
                len(results['self_approved']),
                len(results['al_c_costed']),
                excluded,
            ]
        }
        df_breakdown = pd.DataFrame(breakdown_data)

        table_rows = ""
        for _, r in df_breakdown.iterrows():
            table_rows += f"<tr><td style='padding:0.4rem 0.8rem;'>{r['Category']}</td><td style='padding:0.4rem 0.8rem;text-align:right;font-family:Space Grotesk,sans-serif;font-weight:600;color:#00c4b4;'>{r['Count']}</td></tr>"

        st.markdown(f"""
        <table style='width:100%;border-collapse:collapse;background:#0f0f0f;border-radius:10px;overflow:hidden;font-size:0.9rem;'>
            <thead>
                <tr style='background:#1a1a2e;'>
                    <th style='padding:0.5rem 0.8rem;text-align:left;color:#888;font-weight:500;font-size:0.75rem;text-transform:uppercase;letter-spacing:0.08em;'>Category</th>
                    <th style='padding:0.5rem 0.8rem;text-align:right;color:#888;font-weight:500;font-size:0.75rem;text-transform:uppercase;letter-spacing:0.08em;'>Count</th>
                </tr>
            </thead>
            <tbody>{table_rows}</tbody>
        </table>
        """, unsafe_allow_html=True)

        # ── Flag previews
        if len(results['self_approved']) > 0:
            st.markdown('<div style="margin-top:1.2rem;"></div>', unsafe_allow_html=True)
            with st.expander(f"⚠️ Self-Approved Timesheets ({len(results['self_approved'])} rows)", expanded=False):
                preview_cols = ['Employee Id', 'First Name', 'Surname', 'Timesheet Id', 'Status', 'Location', 'Reviewed By']
                available = [c for c in preview_cols if c in results['self_approved'].columns]
                st.markdown(results['self_approved'][available].to_html(index=False, border=0, classes=''), unsafe_allow_html=True)

        if len(results['al_c_costed']) > 0:
            with st.expander(f"⚠️ Approved Annual Leave on C Locations ({len(results['al_c_costed'])} rows)", expanded=False):
                preview_cols = ['Employee Id', 'First Name', 'Surname', 'Timesheet Id', 'Status', 'Work Type', 'Location']
                available = [c for c in preview_cols if c in results['al_c_costed'].columns]
                st.markdown(results['al_c_costed'][available].to_html(index=False, border=0, classes=''), unsafe_allow_html=True)

        st.markdown('<div class="divider"></div>', unsafe_allow_html=True)

        # ── Generate & Download
        st.markdown("### Step 3 — Download Report")
        output_bytes = build_output_excel(df_for_output, results)

        st.download_button(
            label="⬇️  Download Output File",
            data=output_bytes,
            file_name=output_filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        st.markdown(f"""
        <div class="success-box">
            <strong style="color:#00c4b4;">✓ Report ready</strong><br>
            <span style="color:#aaa;font-size:0.88rem;">
                File: <code style="color:#00c4b4;">{output_filename}</code><br>
                {len(results['approved_non_c'])} approved non-C rows &nbsp;|&nbsp;
                {len(results['unapproved_allocated'])} unapproved allocated rows &nbsp;|&nbsp;
                {flagged} flagged rows &nbsp;|&nbsp;
                {excluded} excluded
            </span>
        </div>
        """, unsafe_allow_html=True)

    except Exception as e:
        st.markdown(f'<div class="error-box"><strong>Error processing file:</strong><br>{str(e)}</div>', unsafe_allow_html=True)
        st.exception(e)

else:
    st.markdown("""
    <div style="background:#0f0f0f;border:1px dashed #1a1a2e;border-radius:12px;padding:2rem;text-align:center;margin-top:1rem;">
        <p style="color:#555;font-size:0.95rem;margin:0;">Upload a Keypay timesheet export above to begin</p>
        <p style="color:#333;font-size:0.8rem;margin-top:0.5rem;">Expected sheet name: <code style="color:#444;">All Timesheets</code></p>
    </div>
    """, unsafe_allow_html=True)
